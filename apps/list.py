from flask import Blueprint, render_template, request, jsonify, url_for, redirect, send_file
from flask.views import MethodView
from flask_login import login_required, current_user
from sqlalchemy import desc, func
from decimal import Decimal
from models import Forms  # Убедитесь, что импортируете правильные модели
import random
from openpyxl import load_workbook
from functions import random_names
from io import BytesIO
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

class ListView(MethodView):
    def __init__(self):
        pass
    
    decorators = [login_required]

    def get(self):
        date_param = request.args.get('date')  # Получите значение параметра "date"
        city_param = request.args.get('where_from')  # Получите значение параметра "city"

        data = Forms.query.filter_by(date=date_param, where_from=city_param).order_by(desc(Forms.number)).all()
        list_delivery = [data for data in data if data.address]
        data_dict = {
            'GEL': {'paid': 0, 'card': 0, 'not_paid': 0},
            'RUB': {'paid': 0, 'card': 0, 'not_paid': 0},
            'USD': {'paid': 0, 'card': 0, 'not_paid': 0},
            'EUR': {'paid': 0, 'card': 0, 'not_paid': 0}
        }
        total_weight = 0
        for record in data:
            try:
                weights_data = record.weights  # Получение значения поля weights из записи
                numbers = [Decimal(num) for num in weights_data.split()]
                total_weight += sum(numbers)  # Суммирование чисел
            except Exception as e:
                total_weight = ' [ ERROR !! ] '

        for item in data:
            currency = item.currency
            payment_status = item.payment_status
            price = int(item.cost)

            data_dict[currency][payment_status] += price

        gel_paid = data_dict['GEL']['paid']
        gel_card = data_dict['GEL']['card']
        gel_not_paid = data_dict['GEL']['not_paid']
        
        rub_paid = data_dict['RUB']['paid']
        rub_card = data_dict['RUB']['card']
        rub_not_paid = data_dict['RUB']['not_paid']

        usd_paid = data_dict['USD']['paid']
        usd_card = data_dict['USD']['card']
        usd_not_paid = data_dict['USD']['not_paid']

        eur_paid = data_dict['EUR']['paid']
        eur_card = data_dict['EUR']['card']
        eur_not_paid = data_dict['EUR']['not_paid']

        return render_template('list.html', data=data, gel_paid=gel_paid, gel_card=gel_card, gel_not_paid=gel_not_paid,
                               rub_paid=rub_paid, rub_card=rub_card, rub_not_paid=rub_not_paid,
                               usd_paid=usd_paid, usd_card=usd_card, usd_not_paid=usd_not_paid,
                               eur_paid=eur_paid, eur_card=eur_card, eur_not_paid=eur_not_paid,
                               total_weight=total_weight, city_param=city_param, date_param=date_param,
                               list_delivery=list_delivery)
    


class AddParcelToList(MethodView):
    decorators = [login_required]

    def __init__(self, db):
        self.db = db

    def post(self):
        access = ['admin', 'Moscow', 'SPB']
        if current_user.role not in access:
            return jsonify({'error': 'User does not have access'}), 403

        try:
            data = request.form.to_dict()
            highest_number = self.db.session.query(func.max(Forms.number)).filter(Forms.date == data['date'],
                                                                        Forms.where_from == data['where_from']).scalar()
            if highest_number is not None:
                new_number = highest_number + 1
            else:
                new_number = 1
            # Получите данные из JSON-запроса
            if data['cost'] == '':
                cost = 0
            else:
                cost = data['cost']

            if data['passport'] == '':
                passport = '---'
            else:
                passport = data['passport']
        
            new_parcel = Forms(
                number=new_number,
                date=data['date'],
                sender_fio=data['sender_fl'].upper(),
                sender_phone=data['sender_phone'],
                recipient_fio=data['recipient_fl'].upper(),
                recipient_phone=data['recipient_phone'],
                passport=passport,
                sender_passport=data['sender_passport'],
                city=data['city'],
                company_comment=data['company_comment'],
                comment=data['comment'],
                price=int(cost),
                weights=data['weights'],
                cost=int(data.get('payment', 0)),
                payment_status=data['payment_status'],
                currency=data['payment_currency'],
                where_from=data['where_from'],
                address=data['address'],
                pdf_adress=data['pdf_adress']

            )

            self.db.session.add(new_parcel)
            self.db.session.commit()

            # Обновите имя эндпоинта на правильное, если требуется
            return redirect(url_for('list', date=data['date'], where_from=data['where_from']))

        except Exception as e:
            # В случае ошибки верните сообщение об ошибке
            response = {"error": str(e)}
            return jsonify(response), 500


class DownloadManifesto(MethodView):
    decorators = [login_required]

    def __init__(self, db):
        self.db = db

    def get(self):
        date = request.args.get('date')
        where_from = request.args.get('where_from')

        # Фильтруем записи в таблице Forms
        filtered_forms = Forms.query.filter(
            Forms.date == date,
            Forms.where_from == where_from,
            Forms.added_to_the_manifest == 'no'
        ).all()

        # Загрузка существующего Excel-файла
        try:
            wb = load_workbook('documents/Sample-Form.xlsx')  # Замените 'Sample-Form.xlsx' на имя вашего существующего файла
            ws = wb.active
        except FileNotFoundError:
            # Обработайте ситуацию, когда файл не найден
            return jsonify({"error": "Sample-Form.xlsx not found"}), 404

        # Обработка данных и запись в Excel
        row_num = ws.max_row + 1  # начинаем с новой строки

        for form in filtered_forms:
            weights = [float(weight) for weight in form.weights.split()]
            price_chance = [15, 20, 25, 10]
            count = 0
            price = random.choice(price_chance)
            vl = 'USD'
            if form.sender_fio:
                if form.sender_fio == 'DAMIR':
                    s_n = random_names()
                else:
                    s_n = form.sender_fio
            else:
                s_n = random_names()
            purc_count = len(weights)
            
            for weight in weights:
                if purc_count != 1:
                    count += 1
                    if where_from == 'Москва':
                        number = f'{form.city}     {form.number}/{count}'
                    else:
                        number = f'{form.city}    0{form.number}/{count}'
                else:
                    if where_from == 'Москва':
                        number = f'{form.city}     {form.number}'
                    else:
                        number = f'{form.city}    0{form.number}'
                # Добавляем данные в соответствующие столбцы
                ws.cell(row=row_num, column=1, value=s_n.split()[0])  # Имя отправителя
                ws.cell(row=row_num, column=2, value=s_n.split()[-1])  # Фамилия отправителя
                ws.cell(row=row_num, column=3, value='Russian Federation')
                if where_from == 'Москва':
                    ws.cell(row=row_num, column=4, value='MOSCOW')
                else:
                    ws.cell(row=row_num, column=4, value='S.P.B')
                ws.cell(row=row_num, column=5, value=form.recipient_fio.split()[0])  # Имя получателя
                ws.cell(row=row_num, column=6, value=form.recipient_fio.split()[-1])  # Фамилия получателя
                ws.cell(row=row_num, column=7, value=form.passport)
                ws.cell(row=row_num, column=8, value='Georgia')
                ws.cell(row=row_num, column=9, value=number)
                ws.cell(row=row_num, column=10, value=form.city)
                ws.cell(row=row_num, column=11, value=form.recipient_phone)
                ws.cell(row=row_num, column=12, value=price)
                ws.cell(row=row_num, column=13, value=vl)
                ws.cell(row=row_num, column=14, value=weight)  # Значение веса

                row_num += 1  # Переходим к следующей строке
            form.added_to_the_manifest = 'yes'
            self.db.session.commit()

       # Сохраняем изменения в оперативной памяти (BytesIO)
        output = BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)  # Возвращаемся в начало буфера

        # Возврат файла для скачивания
        return send_file(output, as_attachment=True, download_name='manifest.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



class CheckPassport(MethodView):
    decorators = [login_required]

    def __init__(self, db):
        self.db = db

    def post(self):

        passport = request.json.get('passport')
        records = Forms.query.filter_by(passport=passport).all()
        
        # Преобразуем дату в объект datetime и сортируем
        sorted_records = sorted(
            records,
            key=lambda x: datetime.strptime(x.date, '%d-%m-%Y'),
            reverse=True
        )
        record = sorted_records[0] if sorted_records else None

        if record:
            return jsonify({
                "found": True,
                "sender_fio": record.sender_fio,
                "sender_phone": record.sender_phone,
                "recipient_fio": record.recipient_fio,
                "recipient_phone": record.recipient_phone,
                "city": record.city,
                "sender_passport": record.sender_passport
                
            })
        else:
            return jsonify({"found": False})




class DownloadInfoExcel(MethodView):
    decorators = [login_required]

    def __init__(self, db):
        self.db = db

    def get(self):
        date = request.args.get('date')
        city = request.args.get('city')

        if not date or not city:
            return jsonify({'error': 'Missing parameters'}), 400

        forms = Forms.query.filter_by(date=date, where_from=city).all()
        if not forms:
            return jsonify({'message': 'Нет данных для указанных параметров'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = 'Info'

        # Стили
        gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        thin = Side(border_style="thin", color="000000")
        thick = Side(border_style="thick", color="000000")
        font = Font(size=14)
        alignment = Alignment(wrap_text=True, vertical='top')

        # Заголовки
        headers = ['Номер посылки', 'Информация отправителя', 'Информация получателя', 'Опись', 'Вес']
        for col_num, title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=title)
            cell.font = font
            cell.alignment = alignment

        current_row = 2

        for index, form in enumerate(forms):
            fill = gray_fill if index % 2 == 0 else white_fill

            ws[f'A{current_row}'] = form.number
            ws[f'B{current_row}'] = form.sender_fio or ''
            ws[f'B{current_row + 1}'] = form.sender_phone or ''
            ws[f'B{current_row + 2}'] = form.sender_passport or ''

            ws[f'C{current_row}'] = form.recipient_fio or ''
            ws[f'C{current_row + 1}'] = form.recipient_phone or ''
            ws[f'C{current_row + 2}'] = form.passport or ''

            comments = [c.strip() for c in (form.comment or '').split(',') if c.strip()]
            for i, comment in enumerate(comments):
                ws[f'D{current_row + i}'] = comment

            weights = re.findall(r'\d+(?:\.\d+)?', form.weights or '')
            weight_sum = sum([float(w) for w in weights])
            ws[f'E{current_row}'] = weight_sum

            used_rows = max(3, len(comments))
            row_start = current_row
            row_end = current_row + used_rows - 1

            for row in range(row_start, row_end + 1):
                for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E'], 1):
                    cell = ws[f'{col_letter}{row}']
                    cell.fill = fill
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = Border(
                        left=thick if col_letter == 'A' else thin,
                        right=thick if col_letter == 'E' else thin,
                        top=thick if row == row_start else thin,
                        bottom=thick if row == row_end else thin
                    )

            current_row += used_rows + 1

        # Автоматическая ширина столбцов
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2  # запас

        # Автоматическая высота строк
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and '\n' in str(cell.value):
                    cell.alignment = Alignment(wrap_text=True)

        # Сохраняем
        output = BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='info_data.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    


def register_list_routes(app, db):
    app.add_url_rule('/list', view_func=ListView.as_view('list'))
    app.add_url_rule('/add_parcell_to_list', view_func=AddParcelToList.as_view('add_parcell_to_list', db=db))
    app.add_url_rule('/download_manifesto', view_func=DownloadManifesto.as_view('download_manifesto', db=db))
    app.add_url_rule('/check_passport', view_func=CheckPassport.as_view('check_passport', db=db))
    app.add_url_rule('/download_info_excel', view_func=DownloadInfoExcel.as_view('download_info_excel', db=db))
