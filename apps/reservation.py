from flask import request, flash, redirect, url_for, render_template, jsonify
from flask_login import login_required, current_user
from flask.views import MethodView
# Импортируйте функцию get_reservation_data из вашего модуля
from functions import get_reservation_data, process_payment, save_booking_to_db, get_existing_booking, update_booking, apply_styles_to_cell  # Замените 'your_module' на имя вашего модуля
# Импортируйте модель Booking, если требуется
from models import Booking  # Замените 'your_module' на имя вашего модуля
import openpyxl
from io import BytesIO
from flask import send_file
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
import re
import qrcode
import os
import json
from datetime import datetime
from openpyxl.drawing.image import Image

BLACKLIST_PATH = os.path.join(os.getcwd(), "documents", "reservation_blacklist.json")


def ensure_blacklist_file():
    os.makedirs(os.path.dirname(BLACKLIST_PATH), exist_ok=True)
    if not os.path.exists(BLACKLIST_PATH):
        with open(BLACKLIST_PATH, "w", encoding="utf-8") as f:
            json.dump([], f, ensure_ascii=False, indent=2)


def normalize_passport(value):
    return re.sub(r"\s+", "", value or "").upper()


def normalize_phone(value):
    return re.sub(r"\D", "", value or "")


def load_reservation_blacklist():
    ensure_blacklist_file()
    try:
        with open(BLACKLIST_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError):
        return []

    return data if isinstance(data, list) else []


def save_reservation_blacklist(entries):
    ensure_blacklist_file()
    with open(BLACKLIST_PATH, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)


def find_blacklist_match(phone, passport):
    entries = load_reservation_blacklist()
    passport_key = normalize_passport(passport)
    phone_key = normalize_phone(phone)

    if passport_key:
        for entry in entries:
            if normalize_passport(entry.get("passport")) == passport_key:
                return entry, "passport"

    if phone_key:
        for entry in entries:
            if normalize_phone(entry.get("phone")) == phone_key:
                return entry, "phone"

    return None, None

class ReservationView(MethodView):
    def __init__(self):
        pass


    decorators = [login_required]

    def get(self):
        access = ['admin', 'Tbilisi', 'Moscow']
        if current_user.role not in access:
            flash('თქვენ არ გაქვთ წვდომა ამ გვერდზე', 'error')
            return redirect(url_for('index'))

        selected_date = request.args.get('date')
        reis = request.args.get('route')
        reservation_data = get_reservation_data(selected_date, reis, 55)
        return render_template('reservation.html', **reservation_data)





class SaveDataView(MethodView):
    def __init__(self, db):
        self.db = db


    decorators = [login_required]

    def post(self):
        # Получаем данные из запроса
        selected_date = request.form.get('selected_date')
        seat_number = request.form.get('seat_number')
        flname = request.form.get('flname').upper()
        gender = request.form.get('gender')
        phone = request.form.get('phone')
        pasport = request.form.get('pasport')
        comment = request.form.get('comment')
        payment = request.form.get('payment')
        fwc = request.form.get('reis')
        destination = request.form.get('destination')
        pay = request.form.get('payment_method')
        pay_method = request.form.get('payment_method_card')
        date_of_birth = request.form.get('date_of_birth')

        # Обработка оплаты
        payment = process_payment(payment, pay, pay_method)

        # Проверяем, существует ли уже запись с таким flname
        existing_booking = Booking.query.filter_by(flname=flname, data=selected_date, fwc=fwc).first()
        if existing_booking:
            return jsonify({
                'success': False,
                'message': 'ამ სახელსა და გვარზე ადგილი უკვე დაჯავშნილია!!!!',
                'message_ru': 'Такой человек уже забронирован на этот рейс'
            })
        else:
            # Сохранение данных в базе данных
            save_booking_to_db(self.db, selected_date, seat_number, flname, gender, phone, pasport, comment, payment, fwc, destination, date_of_birth)

            return jsonify({"success": True, "message": "Данные успешно сохранены"})




class EditBookingView(MethodView):
    def __init__(self, db):
        self.db = db


    decorators = [login_required]

    def post(self):
        # Получаем данные из запроса
        seat_number = request.form.get('seat_number')
        old_seat_number = request.form.get('old_seat_number')  # Получаем старое место
        gender = request.form.get('gender')
        flname = request.form.get('flname').upper()
        phone = request.form.get('phone')
        pasport = request.form.get('pasport')
        comment = request.form.get('comment')
        payment = request.form.get('payment')
        destination = request.form.get('destination')
        reis = request.form.get('reis')
        selected_date = request.form.get('selected_date')
        pay = request.form.get('payment_method')
        pay_method = request.form.get('payment_method_card')
        date_of_birth = request.form.get('date_of_birth')

        # Обработка оплаты
        payment = process_payment(payment, pay, pay_method)

        # Получаем существующие бронирования
        existing_booking = get_existing_booking(reis, selected_date, old_seat_number)
        existing_booking_old = get_existing_booking(reis, selected_date, seat_number)

        if seat_number == old_seat_number:
            if existing_booking_old:
                # Обновление бронирования
                update_booking(self.db, existing_booking_old, gender, flname, phone, pasport, payment, destination, comment, seat_number, date_of_birth)
                return jsonify({'success': True, 'message': 'Бронирование обновлено успешно'})

        if existing_booking:
            return jsonify({'success': False, 'message': 'Садовое место уже занято'})
        elif existing_booking is None:
            # Обновление бронирования
            update_booking(self.db, existing_booking_old, gender, flname, phone, pasport, payment, destination, comment, old_seat_number, date_of_birth)
            return jsonify({'success': True, 'message': 'Бронирование обновлено успешно'})



class BookingDeleteView(MethodView):
    def __init__(self, db):
        self.db = db


    decorators = [login_required]

    def post(self):
        # Получаем данные из запроса
        seat_number = request.form.get('s_n')
        reis = request.form.get('reis')
        selected_date = request.form.get('selected_date')
        
        # Ищем существующее бронирование
        existing_booking = Booking.query.filter_by(fwc=reis, data=selected_date, position=seat_number).first()
        
        if existing_booking:
            # Удаляем запись из базы данных
            self.db.session.delete(existing_booking)
            self.db.session.commit()
            return jsonify({'success': True, 'message': 'Бронирование успешно удалено'})
        
        return jsonify({'success': False, 'message': 'Бронирование не найдено'})



class DownloadVedView(MethodView):
    def __init__(self, db):
        self.db = db

    decorators = [login_required]

    def post(self):
        if current_user.role == 'Moscow':
            destination_address = 'Тбилиси'
        elif current_user.role == 'Tbilisi':
            destination_address = 'Москва'
        else:
            destination_address = ''
        
        reis = request.form.get('reis')
        selected_date = request.form.get('selected_date')
        filtered_data = Booking.query.filter(
            Booking.fwc == reis,
            Booking.data == selected_date,
            Booking.action == 'yes'
        ).all()

        workbook = openpyxl.load_workbook('documents/vedom.xlsx')
        sheet = workbook.active

        start_row = 7
        col_letters = ['B', 'D', 'E', 'F', 'G']

        # Устанавливаем стиль шрифта по умолчанию с размером 10
        font = Font(size=10)
        for row_idx, data in enumerate(filtered_data, start=start_row):
            for col_idx, col_letter in enumerate(col_letters):
                cell = sheet[f'{col_letter}{row_idx}']
                if col_letter == 'B':
                    cell.value = data.date_of_birth
                elif col_letter == 'D':
                    cell.value = data.pasport
                elif col_letter == 'E':
                    cell.value = data.flname
                elif col_letter == 'F':
                    cell.value = data.position
                elif col_letter == 'G':
                    cell.value = destination_address
                cell.font = font  # Применяем стиль шрифта к ячейке

        # Сохранение файла в оперативную память
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        workbook.close()

        # Отправляем файл для скачивания
        return send_file(
            output,
            as_attachment=True,
            download_name=f'Ведомость от {selected_date}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )




class GenerateTicketView(MethodView):
    def __init__(self, db):
        self.db = db

    decorators = [login_required]

    def post(self):
        try:
            s_n = request.form.get('s_n')
            reis = request.form.get('reis')
            selected_date = request.form.get('selected_date')

            # Загружаем шаблон Excel-файла
            template_path = 'documents/ticket.xlsx'
            workbook = load_workbook(template_path)
            sheet = workbook.active

            # По вашим указаниям, заполняем ячейки данными
            booking = Booking.query.filter(
                Booking.fwc == reis,
                Booking.data == selected_date,
                Booking.position == s_n
            ).first()


            if booking:
                booking.action = 'yes'
                self.db.session.commit()  # Сохранение изменений в базе данных
                bold_font = Font(bold=True)

                apply_styles_to_cell(sheet, 'G5', booking.flname)
                
                apply_styles_to_cell(sheet, 'S5', booking.pasport)

                apply_styles_to_cell(sheet, 'N8', booking.data)

                apply_styles_to_cell(sheet, 'W8', booking.destination)

                apply_styles_to_cell(sheet, 'A12', booking.position)

                if current_user.role == 'Moscow':
                    sheet['A15'] = '10 : 00'
                else:
                    sheet['A15'] = '10 : 00'
                sheet['A15'].alignment = Alignment(horizontal='center', vertical='center')
                sheet['A15'].font = bold_font


                if booking.gender == 'male':
                    sheet['G8'] = 'М / მმ'
                elif booking.gender == 'female':
                    sheet['G8'] = 'Ж / მდ'
                sheet['G8'].alignment = Alignment(horizontal='center', vertical='center')
                sheet['G8'].font = bold_font

                # Обработка оплаты
                payment_value = booking.payment
                payment_currency = payment_value[-3:]
                if payment_currency == 'RUB' and current_user.role == 'Moscow':
                    sheet['A7'] = '₽'
                    sheet['A7'].alignment = Alignment(horizontal='center', vertical='center')
                    sheet['A7'].font = bold_font
                    sheet['D7'] = re.sub(r'\D', '', payment_value)  # Убираем первую букву
                    sheet['D7'].alignment = Alignment(horizontal='center', vertical='center')
                    sheet['D7'].font = bold_font
                elif payment_currency == 'GEL':
                    sheet['A7'] = '₾'

                    sheet['A7'].alignment = Alignment(horizontal='center', vertical='center')
                    sheet['A7'].font = bold_font
                    sheet['D7'] = re.sub(r'\D', '', payment_value)  # Убираем первую букву
                    sheet['D7'].alignment = Alignment(horizontal='center', vertical='center')
                    sheet['D7'].font = bold_font
                else:
                    sheet['A7'] = '₾'
                    sheet['A7'].alignment = Alignment(horizontal='center', vertical='center')
                    sheet['A7'].font = bold_font
                    sheet['D7'] = '200'
                    sheet['D7'].alignment = Alignment(horizontal='center', vertical='center')
                    sheet['D7'].font = bold_font
                    

                timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                data = f"კომპანია VIP-TOUR, მისამართი: გულიას ქუჩა №5. ბილეთის მიღების დრო {timestamp}. ბოლო გაჩერება-ოფისი: იუჟნაპორტოვაიას ქუჩა, სახლი 7, შენობა 25.\nКомпания VIP-TOUR, Адрес:  ул.Гулия №5.  / Время получения билета {timestamp}.  / конечная остановка-офис: ул.Южнопоровая дом 7, строение 25."
                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=6.8,
                    border=0,
                )
                qr.add_data(data)
                qr.make(fit=True)
                img = qr.make_image(fill_color="black", back_color="white")
                img.save("qr_code.png")

                # Вставьте QR-код в Excel
                img = Image("qr_code.png")
                img.width = img.width // 3.2  # Уменьшите размер изображения, если необходимо
                img.height = img.height // 3.2
                sheet.add_image(img, "X10")


                # Сохраняем Excel в память
                output = BytesIO()
                workbook.save(output)
                output.seek(0)
                
                # Отправляем файл для скачивания
                return send_file(output, as_attachment=True, download_name=f'ticket-{s_n}_for{selected_date}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            else:
                return jsonify({'success': False, 'message': 'Данные не найдены'})

        except Exception as e:
            return jsonify({'success': False, 'message': str(e)})


class UserCheckWithPhone(MethodView):
    decorators = [login_required]

    def __init__(self, db):
        self.db = db

    def post(self):
        data = request.get_json()
        phone = data.get("phone")

        if not phone:
            return jsonify({"success": False, "error": "ტელეფონი არ არის მითითებული"}), 400

        # Ищем запись с указанным телефоном, начиная с самой новой (по id или дате)
        user = Booking.query.filter_by(phone=phone).order_by(Booking.id.desc()).first()

        if user:
            return jsonify({
                "success": True,
                "flname": user.flname,
                "gender": user.gender,
                "pasport": user.pasport,
                "date_of_birth": user.date_of_birth
            })
        else:
            return jsonify({"success": False})


class UserCheckWithPassport(MethodView):
    decorators = [login_required]

    def __init__(self, db):
        self.db = db

    def post(self):
        data = request.get_json()
        passport = data.get("pasport")

        if not passport:
            return jsonify({"success": False, "error": "საბუთი არ არის მითითებული"}), 400

        # Ищем запись с указанным телефоном, начиная с самой новой (по id или дате)
        user = Booking.query.filter_by(pasport=passport).order_by(Booking.id.desc()).first()

        if user:
            return jsonify({
                "success": True,
                "flname": user.flname,
                "gender": user.gender,
                "phone": user.phone,
                "date_of_birth": user.date_of_birth
            })
        else:
            return jsonify({"success": False})


class ReservationBlacklistCheckView(MethodView):
    decorators = [login_required]

    def post(self):
        data = request.get_json() or {}
        phone = data.get("phone", "")
        passport = data.get("passport", "")

        entry, match_by = find_blacklist_match(phone, passport)

        if not entry:
            return jsonify({"found": False})

        return jsonify({
            "found": True,
            "match_by": match_by,
            "phone": entry.get("phone", ""),
            "passport": entry.get("passport", ""),
            "reason": entry.get("reason", "")
        })


class ReservationBlacklistAddView(MethodView):
    decorators = [login_required]

    def post(self):
        data = request.get_json() or {}
        phone = (data.get("phone") or "").strip()
        passport = (data.get("passport") or "").strip()
        reason = (data.get("reason") or "").strip()

        if not passport:
            return jsonify({
                "success": False,
                "message": "Укажите паспорт",
                "message_ru": "Укажите паспорт",
                "message_ge": "მიუთითეთ პასპორტი"
            }), 400

        if not reason:
            return jsonify({
                "success": False,
                "message": "Укажите причину",
                "message_ru": "Укажите причину",
                "message_ge": "მიუთითეთ მიზეზი"
            }), 400

        entries = load_reservation_blacklist()
        passport_key = normalize_passport(passport)
        phone_key = normalize_phone(phone)

        for entry in entries:
            if normalize_passport(entry.get("passport")) == passport_key:
                return jsonify({
                    "success": False,
                    "message": "Такой паспорт уже есть в черном списке",
                    "message_ru": "Такой паспорт уже есть в черном списке",
                    "message_ge": "ასეთი პასპორტი უკვე არის შავ სიაში"
                }), 409

            if phone_key and normalize_phone(entry.get("phone")) == phone_key:
                return jsonify({
                    "success": False,
                    "message": "Такой телефон уже есть в черном списке",
                    "message_ru": "Такой телефон уже есть в черном списке",
                    "message_ge": "ასეთი ტელეფონი უკვე არის შავ სიაში"
                }), 409

        entries.append({
            "passport": passport,
            "phone": phone,
            "reason": reason
        })
        save_reservation_blacklist(entries)

        return jsonify({
            "success": True,
            "message": "Запись добавлена в черный список",
            "message_ru": "Запись добавлена в черный список",
            "message_ge": "ჩანაწერი დაემატა შავ სიაში"
        })






def register_reservation_routes(app, db):
    app.add_url_rule('/reservation', view_func=ReservationView.as_view('reservation'))
    app.add_url_rule('/save_data', view_func=SaveDataView.as_view('save_data', db=db))
    app.add_url_rule('/edit_booking', view_func=EditBookingView.as_view('edit_booking', db=db))
    app.add_url_rule('/booking_del', view_func=BookingDeleteView.as_view('booking_del', db=db))
    app.add_url_rule('/download_ved', view_func=DownloadVedView.as_view('download_ved', db=db))
    app.add_url_rule('/ticket', view_func=GenerateTicketView.as_view('generate_ticket', db=db))
    app.add_url_rule('/check_user_with_phone', view_func=UserCheckWithPhone.as_view('check_user_with_phone', db=db))
    app.add_url_rule('/check_user_with_passport', view_func=UserCheckWithPassport.as_view('check_user_with_passport', db=db))
    app.add_url_rule('/reservation_blacklist_check', view_func=ReservationBlacklistCheckView.as_view('reservation_blacklist_check'))
    app.add_url_rule('/reservation_blacklist_add', view_func=ReservationBlacklistAddView.as_view('reservation_blacklist_add'))
