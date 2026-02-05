from flask.views import MethodView
from flask_login import login_required, current_user
from flask import Blueprint, render_template, request, jsonify, url_for, redirect, send_file
import os
import json
from models import Shipments;
from datetime import datetime;
from decimal import Decimal
from io import BytesIO
import random
from openpyxl import load_workbook
from models import Forms
from functions import random_names
from helper.shipments_helper import weight_list, extract_inventory_names, split_fio


class ListView(MethodView):
    decorators = [login_required]

    def get(self):
        date_param = datetime.strptime(request.args.get('date'), "%d-%m-%Y").date() 
        city_param = request.args.get('where_from')  # –ü–æ–ª—É—á–∏—Ç–µ –∑–Ω–∞—á–µ–Ω–∏–µ –ø–∞—Ä–∞–º–µ—Ç—Ä–∞ "city"

        # 1. –ü–æ–ª—É—á–∞–µ–º –≤—Å–µ –ø–æ—Å—ã–ª–∫–∏ –∏–∑ –ë–î
        shipments = (Shipments.query.filter_by(send_date=date_param,city_from=city_param).order_by(Shipments.id.desc()).all())

        # 2. –ó–∞–≥—Ä—É–∂–∞–µ–º inventory (–∫–∞–∫ —É —Ç–µ–±—è –±—ã–ª–æ)
        json_path = os.path.join(os.getcwd(), "documents", "inventory.json")
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                inventory_items = json.load(f)
        except FileNotFoundError:
            inventory_items = []

        # 3. –ü–µ—Ä–µ–¥–∞—ë–º –í–°–Å –≤ —à–∞–±–ª–æ–Ω
        return render_template(
            "shipments.html",
            shipments=shipments,
            inventory=inventory_items
        )

# POST –æ–±—Ä–∞–±–æ—Ç—á–∏–∫ –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –¥–∞–Ω–Ω—ã—Ö –∏–∑ –º–æ–¥–∞–ª–∫–∏
class ShipmentSubmitView(MethodView):
    decorators = [login_required]

    def __init__(self, db):
        self.db = db

    def post(self):
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "message": "–ù–µ—Ç –¥–∞–Ω–Ω—ã—Ö"}), 400

        # –ø—Ä–æ–≤–µ—Ä—è–µ–º, —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –∏–ª–∏ –Ω–æ–≤–æ–µ –¥–æ–±–∞–≤–ª–µ–Ω–∏–µ
        shipment_id = data.get("shipmentId")  # <-- –ø–µ—Ä–µ–¥–∞–µ–º –∏–∑ JS –ø—Ä–∏ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–∏

        total_weight, parcels_count = weight_list(data)

        # üîπ –î–∞–Ω–Ω—ã–µ –∏–∑ URL
        date_param = datetime.strptime(data.get("date"), "%d-%m-%Y").date()
        where_from_param = data.get("where_from")

        # –µ—Å–ª–∏ –Ω–æ–≤–æ–µ –¥–æ–±–∞–≤–ª–µ–Ω–∏–µ, —Å—á–∏—Ç–∞–µ–º –Ω–æ–º–µ—Ä –ø–æ—Å—ã–ª–∫–∏
        if not shipment_id:
            last_shipment = (
                Shipments.query
                .filter(Shipments.send_date == date_param, Shipments.city_from == where_from_param)
                .order_by(Shipments.shipment_number.desc())
                .first()
            )
            shipment_number = 1 if last_shipment is None else last_shipment.shipment_number + 1

        shared_recipient = data.get("sharedRecipient")
        if shared_recipient:
            payment_amount = 0
            payment_status = "+"
            sequence = 1
        else:
            payment_amount = data.get("paymentAmount", 0)
            payment_status = data.get("paymentStatus", "")
            sequence = 0

        inventory = data.get("inventory", [])
        clean_inventory = [item.replace("√ó", "").strip() for item in inventory]
        
        inventory_names = extract_inventory_names(clean_inventory)
        storage = InventoryStorage()
        storage.add_new_items(inventory_names)

        try:
            if shipment_id:  # —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ
                parcel = Shipments.query.get(shipment_id)
                if not parcel:
                    return jsonify({"success": False, "message": "–ü–æ—Å—ã–ª–∫–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω–∞"}), 404
            else:  # –Ω–æ–≤–æ–µ –¥–æ–±–∞–≤–ª–µ–Ω–∏–µ
                parcel = Shipments(
                    shipment_number=shipment_number,
                    city_from=where_from_param,
                    send_date=date_param,
                    order_date=datetime.now(),
                )
                self.db.session.add(parcel)

            # ‚úÖ –û–±—â–∏–µ –ø–æ–ª—è (–∏ –¥–ª—è —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è, –∏ –¥–ª—è –Ω–æ–≤–æ–≥–æ)
            parcel.sender_name = data.get("senderName", "")
            parcel.sender_surname = data.get("senderSurname", "")
            parcel.sender_number = data.get("senderPhone", "")

            parcel.recipient_name = data.get("recipientName", "")
            parcel.recipient_surname = data.get("recipientSurname", "")
            parcel.recipient_number = data.get("recipientPhone", "")
            parcel.recipient_passport = data.get("recipientPassport", "")

            parcel.weights = data.get("weightsHidden", "")
            parcel.total_weight = total_weight
            parcel.parcels_count = parcels_count
            parcel.city_to = data.get("parcelCity", "")
            parcel.cargo_cost = data.get("parcelCost", "")
            parcel.address = data.get("parcelAddress", "")

          

            parcel.description = ", ".join(clean_inventory)
            parcel.payment_amount = payment_amount
            parcel.payment_status = payment_status
            parcel.currency = data.get("currency", "")
            parcel.sequence = sequence

            parcel.information_for_office = data.get("informationForOffice", "")

            self.db.session.commit()
            message = "–ü–æ—Å—ã–ª–∫–∞ —É—Å–ø–µ—à–Ω–æ –æ–±–Ω–æ–≤–ª–µ–Ω–∞" if shipment_id else "–ü–æ—Å—ã–ª–∫–∞ —É—Å–ø–µ—à–Ω–æ –¥–æ–±–∞–≤–ª–µ–Ω–∞"
            return jsonify({"success": True, "message": message})

        except Exception as e:
            self.db.session.rollback()
            return jsonify({"success": False, "message": str(e)}), 500







class DownloadManifest(MethodView):
    decorators = [login_required]

    def __init__(self, db):
        self.db = db

    def post(self):   # üî• –í–ê–ñ–ù–û
        data = request.get_json() or {}
        shipment_ids = data.get("shipment_ids", [])

        if not shipment_ids:
            return jsonify({"error": "shipment_ids –ø—É—Å—Ç"}), 400

        # –§–∏–ª—å—Ç—Ä—É–µ–º –∑–∞–ø–∏—Å–∏ –≤ —Ç–∞–±–ª–∏—Ü–µ Shipments
        filtered_forms = Shipments.query.filter(
            Shipments.id.in_(shipment_ids)  # üîπ —Ñ–∏–ª—å—Ç—Ä –ø–æ —Å–ø–∏—Å–∫—É ID
        ).all()

        # –ó–∞–≥—Ä—É–∑–∫–∞ —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–≥–æ Excel-—Ñ–∞–π–ª–∞
        try:
            wb = load_workbook('documents/Sample-Form.xlsx')  # –ó–∞–º–µ–Ω–∏—Ç–µ 'Sample-Form.xlsx' –Ω–∞ –∏–º—è –≤–∞—à–µ–≥–æ —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–≥–æ —Ñ–∞–π–ª–∞
            ws = wb.active
        except FileNotFoundError:
            # –û–±—Ä–∞–±–æ—Ç–∞–π—Ç–µ —Å–∏—Ç—É–∞—Ü–∏—é, –∫–æ–≥–¥–∞ —Ñ–∞–π–ª –Ω–µ –Ω–∞–π–¥–µ–Ω
            return jsonify({"error": "Sample-Form.xlsx not found"}), 404

        # –û–±—Ä–∞–±–æ—Ç–∫–∞ –¥–∞–Ω–Ω—ã—Ö –∏ –∑–∞–ø–∏—Å—å –≤ Excel
        row_num = ws.max_row + 1  # –Ω–∞—á–∏–Ω–∞–µ–º —Å –Ω–æ–≤–æ–π —Å—Ç—Ä–æ–∫–∏

        for form in filtered_forms:
            weights = [float(weight) for weight in form.weights.split()]
            price_chance = [15, 20, 25, 10]
            count = 0
            price = random.choice(price_chance)
            vl = 'USD'
            if form.sender_name:
                if form.sender_name == 'DAMIR':
                    s_n = random_names()
                else:
                    s_n = f'{form.sender_name} {form.sender_surname}'
            else:
                s_n = random_names()
            purc_count = len(weights)
            
            for weight in weights:
                if purc_count != 1:
                    count += 1
                    if form.city_from == '–ú–æ—Å–∫–≤–∞':
                        number = f'{form.city_to}     {form.shipment_number}/{count}'
                    else:
                        number = f'{form.city_to}    0{form.shipment_number}/{count}'
                else:
                    if form.city_from == '–ú–æ—Å–∫–≤–∞':
                        number = f'{form.city_to}     {form.shipment_number}'
                    else:
                        number = f'{form.city_to}    0{form.shipment_number}'
                # –î–æ–±–∞–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –≤ —Å–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤—É—é—â–∏–µ —Å—Ç–æ–ª–±—Ü—ã
                ws.cell(row=row_num, column=1, value=s_n.split()[0])  # –ò–º—è –æ—Ç–ø—Ä–∞–≤–∏—Ç–µ–ª—è
                ws.cell(row=row_num, column=2, value=s_n.split()[-1])  # –§–∞–º–∏–ª–∏—è –æ—Ç–ø—Ä–∞–≤–∏—Ç–µ–ª—è
                ws.cell(row=row_num, column=3, value='Russian Federation')
                if form.city_from == '–ú–æ—Å–∫–≤–∞':
                    ws.cell(row=row_num, column=4, value='MOSCOW')
                else:
                    ws.cell(row=row_num, column=4, value='S.P.B')
                ws.cell(row=row_num, column=5, value=form.recipient_name)  # –ò–º—è –ø–æ–ª—É—á–∞—Ç–µ–ª—è
                ws.cell(row=row_num, column=6, value=form.recipient_surname)  # –§–∞–º–∏–ª–∏—è –ø–æ–ª—É—á–∞—Ç–µ–ª—è
                ws.cell(row=row_num, column=7, value=form.recipient_passport)
                ws.cell(row=row_num, column=8, value='Georgia')
                ws.cell(row=row_num, column=9, value=number)
                ws.cell(row=row_num, column=10, value=form.city_to)
                ws.cell(row=row_num, column=11, value=form.recipient_number)
                ws.cell(row=row_num, column=12, value=price)
                ws.cell(row=row_num, column=13, value=vl)
                ws.cell(row=row_num, column=14, value=weight)  # –ó–Ω–∞—á–µ–Ω–∏–µ –≤–µ—Å–∞

                row_num += 1  # –ü–µ—Ä–µ—Ö–æ–¥–∏–º –∫ —Å–ª–µ–¥—É—é—â–µ–π —Å—Ç—Ä–æ–∫–µ

       # –°–æ—Ö—Ä–∞–Ω—è–µ–º –∏–∑–º–µ–Ω–µ–Ω–∏—è –≤ –æ–ø–µ—Ä–∞—Ç–∏–≤–Ω–æ–π –ø–∞–º—è—Ç–∏ (BytesIO)
        output = BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)  # –í–æ–∑–≤—Ä–∞—â–∞–µ–º—Å—è –≤ –Ω–∞—á–∞–ª–æ –±—É—Ñ–µ—Ä–∞

        # ‚úÖ –û—Ç–º–µ—á–∞–µ–º –ø–æ—Å—ã–ª–∫–∏ –∫–∞–∫ –æ–±—Ä–∞–±–æ—Ç–∞–Ω–Ω—ã–µ
        Shipments.query.filter(
            Shipments.id.in_(shipment_ids)
        ).update(
            {Shipments.is_processed: True},
            synchronize_session=False
        )
        self.db.session.commit()

        # –í–æ–∑–≤—Ä–∞—Ç —Ñ–∞–π–ª–∞ –¥–ª—è —Å–∫–∞—á–∏–≤–∞–Ω–∏—è
        return send_file(output, as_attachment=True, download_name='manifest.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    


class ShipmentDetailView(MethodView):
    decorators = [login_required]

    def __init__(self, db):
        self.db = db

    def get(self, shipment_id):
        shipment = Shipments.query.get_or_404(shipment_id)

        return jsonify({
            "id": shipment.id,

            # –û—Ç–ø—Ä–∞–≤–∏—Ç–µ–ª—å
            "sender_name": shipment.sender_name,
            "sender_surname": shipment.sender_surname,
            "sender_number": shipment.sender_number,

            # –ü–æ–ª—É—á–∞—Ç–µ–ª—å
            "recipient_name": shipment.recipient_name,
            "recipient_surname": shipment.recipient_surname,
            "recipient_number": shipment.recipient_number,
            "recipient_passport": shipment.recipient_passport,

            # –ü–æ—Å—ã–ª–∫–∞
            "weights": shipment.weights,
            "city_to": shipment.city_to,
            "cargo_cost": shipment.cargo_cost,
            "address": shipment.address,

            # –û–ø–∏—Å—å
            "description": shipment.description,
            "information_for_office": shipment.information_for_office,

            # –û–ø–ª–∞—Ç–∞
            "payment_amount": shipment.payment_amount,
            "payment_status": shipment.payment_status,
            "currency": shipment.currency
        })

class InventoryStorage:
    def __init__(self):
        self.json_path = os.path.join(os.getcwd(), "documents", "inventory.json")

    def load(self):
        try:
            with open(self.json_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except FileNotFoundError:
            return []

    def save(self, items):
        with open(self.json_path, "w", encoding="utf-8") as f:
            json.dump(items, f, ensure_ascii=False, indent=2)

    def add_new_items(self, new_items):
        """
        new_items ‚Üí ["–æ–¥–µ–∂–¥–∞", "–æ–±—É–≤—å"]
        """
        existing_items = self.load()
        updated = False

        for item in new_items:
            if item not in existing_items:
                existing_items.append(item)
                updated = True

        if updated:
            self.save(existing_items)



class DownloadInventory(MethodView):
    decorators = [login_required]

    def __init__(self, db):
        self.db = db

    def post(self):
        data = request.get_json() or {}
        shipment_ids = data.get("shipment_ids", [])

        if not shipment_ids:
            return jsonify({"error": "shipment_ids –ø—É—Å—Ç"}), 400

        # –ü–æ–ª—É—á–∞–µ–º –≤—ã–±—Ä–∞–Ω–Ω—ã–µ –ø–æ—Å—ã–ª–∫–∏
        filtered_forms = Shipments.query.filter(
            Shipments.id.in_(shipment_ids)
        ).all()

        # –ó–∞–≥—Ä—É–∂–∞–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–π Excel-—à–∞–±–ª–æ–Ω
        try:
            wb = load_workbook('documents/Sample-List.xlsx')  # –®–∞–±–ª–æ–Ω —Å –∑–∞–≥–æ–ª–æ–≤–∫–∞–º–∏ –≤ –ø–µ—Ä–≤–æ–π —Å—Ç—Ä–æ–∫–µ
            ws = wb.active
        except FileNotFoundError:
            return jsonify({"error": "Sample-List.xlsx not found"}), 404

        # –§–æ—Ä–º–∏—Ä—É–µ–º —Å–ª–æ–≤–∞—Ä—å –¥–ª—è —Å—É–º–º–∏—Ä–æ–≤–∞–Ω–∏—è
        # { "–û–¥–µ–∂–¥–∞": {"weight": 10.5, "count": 2}, ... }
        inventory_dict = {}

        for form in filtered_forms:
            if not form.description:
                continue
            items = form.description.split(",")  # —Ä–∞–∑–¥–µ–ª—è–µ–º –ø–æ –∑–∞–ø—è—Ç—ã–º
            for item in items:
                item = item.strip()
                if not item:
                    continue
                # –†–∞–∑–¥–µ–ª—è–µ–º –Ω–∞–∑–≤–∞–Ω–∏–µ –∏ –≤–µ—Å
                if ":" in item:
                    name, weight = item.split(":")
                    name = name.strip()
                    try:
                        weight = float(weight.strip())
                    except ValueError:
                        weight = 0
                else:
                    name = item
                    weight = 0

                if name in inventory_dict:
                    inventory_dict[name]["weight"] += weight
                    inventory_dict[name]["count"] += 1
                else:
                    inventory_dict[name] = {"weight": weight, "count": 1}

        # –ù–∞—á–∏–Ω–∞–µ–º –∑–∞–ø–∏—Å—å —Å–æ –≤—Ç–æ—Ä–æ–π —Å—Ç—Ä–æ–∫–∏ (1-based indexing)
        row_num = 2

        for name, data_item in inventory_dict.items():
            ws.cell(row=row_num, column=1, value=name)                  # A ‚Äî –Ω–∞–∑–≤–∞–Ω–∏–µ —Ç–æ–≤–∞—Ä–∞
            ws.cell(row=row_num, column=2, value=data_item["weight"])  # B ‚Äî –æ–±—â–∏–π –≤–µ—Å
            ws.cell(row=row_num, column=3, value=data_item["count"])   # C ‚Äî –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ
            row_num += 1

        # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤ –æ–ø–µ—Ä–∞—Ç–∏–≤–Ω–æ–π –ø–∞–º—è—Ç–∏
        output = BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)

        # –û—Ç–¥–∞–µ–º —Ñ–∞–π–ª –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é
        return send_file(
            output,
            as_attachment=True,
            download_name='inventory.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )






class CheckSender(MethodView):
    decorators = [login_required]

    def __init__(self, db):
        self.db = db

    def post(self):
        data = request.get_json()
        recipient_passport = data.get('passport')

        if not recipient_passport:
            return jsonify({"found": False})

        # 1Ô∏è‚É£ –ü—ã—Ç–∞–µ–º—Å—è –Ω–∞–π—Ç–∏ –≤ –ù–û–í–û–ô —Ç–∞–±–ª–∏—Ü–µ
        shipment = (
            Shipments.query
            .filter(Shipments.recipient_passport == recipient_passport)
            .order_by(Shipments.id.desc())
            .first()
        )

        if shipment:
            return jsonify({
                "found": True,
                "source": "new",
                "sender": {
                    "name": shipment.sender_name,
                    "surname": shipment.sender_surname,
                    "phone": shipment.sender_number
                },
                "recipient": {
                    "name": shipment.recipient_name,
                    "surname": shipment.recipient_surname,
                    "phone": shipment.recipient_number,
                    "passport": shipment.recipient_passport
                },
                "address": {
                    "city": shipment.city_to,
                    "street": shipment.address
                }
            })

        # 2Ô∏è‚É£ –ï—Å–ª–∏ –Ω–µ –Ω–∞—à–ª–∏ ‚Äî –∏—â–µ–º –≤ –°–¢–ê–†–û–ô —Ç–∞–±–ª–∏—Ü–µ
        form = (
            Forms.query
            .filter(Forms.passport == recipient_passport)
            .order_by(Forms.id.desc())
            .first()
        )

        if not form:
            return jsonify({"found": False})

        # 3Ô∏è‚É£ –†–∞–∑–±–∏–≤–∞–µ–º –§–ò–û
        sender_name, sender_surname = split_fio(form.sender_fio)
        recipient_name, recipient_surname = split_fio(form.recipient_fio)
        print(sender_name, sender_surname)

        return jsonify({
            "found": True,
            "source": "old",
            "sender": {
                "name": sender_name,
                "surname": sender_surname,
                "phone": form.sender_phone
            },
            "recipient": {
                "name": recipient_name,
                "surname": recipient_surname,
                "phone": form.recipient_phone,
                "passport": form.passport
            },
            "address": {
                "city": form.city,
                "street": form.address
            }
        })






def register_shipments_routes(app, db):
    app.add_url_rule('/shipments', view_func=ListView.as_view('shipments'))

    app.add_url_rule('/shipment_submit', view_func=ShipmentSubmitView.as_view('shipment_submit', db=db))
    

    app.add_url_rule('/download_manifest', view_func=DownloadManifest.as_view('download_manifest', db=db))

    shipment_detail_view = ShipmentDetailView.as_view("shipment_detail", db=db)
    app.add_url_rule("/shipments/<int:shipment_id>", view_func=shipment_detail_view, methods=["GET"])

    app.add_url_rule('/download_inventory', view_func=DownloadInventory.as_view('download_inventory', db=db))

    app.add_url_rule('/check_sender', view_func=CheckSender.as_view('check_sender', db=db))